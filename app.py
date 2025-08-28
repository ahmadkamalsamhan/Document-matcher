import streamlit as st
import pandas as pd
import re
import tempfile
import os
import time
from openpyxl import load_workbook

st.set_page_config(page_title="King Salman Park - Matching & Filter App", layout="wide")
st.title("📊 King Salman Park - Document Processing App")

# -----------------------------
# GLOBAL RESET BUTTON
# -----------------------------
if st.button("🗑 Clear/Reset Entire App"):
    keys_to_clear = ["uploaded_files", "tmp_path", "filter_file"]
    cleared = False
    for key in keys_to_clear:
        if key in st.session_state:
            del st.session_state[key]
            cleared = True
    if cleared:
        st.success("✅ App fully reset. All uploaded files and filters cleared.")
        st.experimental_rerun()
    else:
        st.success("✅ App is already clean. You can continue normally.")

# -----------------------------
# PART 1 - MATCHING
# -----------------------------
st.header("🔹 Part 1: Matching Two Excel Files")

uploaded_files = st.file_uploader(
    "Upload Excel files", type="xlsx", accept_multiple_files=True, key="uploaded_files"
)

if uploaded_files:
    st.subheader("Select files to use for matching")
    selected_files = [f for f in uploaded_files if st.checkbox(f.name, value=True)]
    if len(selected_files) >= 2:
        st.success(f"{len(selected_files)} files selected for matching.")

        df1_columns = pd.read_excel(selected_files[0], nrows=0).columns.tolist()
        df2_columns = pd.read_excel(selected_files[1], nrows=0).columns.tolist()

        st.subheader("Step 1: Select column to match")
        match_col1 = st.selectbox(f"Column from {selected_files[0].name}", df1_columns)
        match_col2 = st.selectbox(f"Column from {selected_files[1].name}", df2_columns)

        st.subheader("Step 2: Select additional columns to include in the result")
        include_cols1 = st.multiselect(f"Columns from {selected_files[0].name}", df1_columns)
        include_cols2 = st.multiselect(f"Columns from {selected_files[1].name}", df2_columns)

        if st.button("Step 3: Start Matching"):
            if not match_col1 or not match_col2:
                st.warning("⚠️ Please select columns to match.")
            elif not include_cols1 and not include_cols2:
                st.warning("⚠️ Please select at least one additional column to include in the result.")
            else:
                st.info("⏳ Matching in progress (exact Colab logic, memory-safe)...")
                try:
                    df1_small = pd.read_excel(selected_files[0], usecols=[match_col1] + include_cols1)
                    df2_small = pd.read_excel(selected_files[1], usecols=[match_col2] + include_cols2)

                    def normalize(text):
                        if pd.isna(text): return ""
                        text = str(text).lower()
                        text = re.sub(r'[^a-z0-9\s]', ' ', text)
                        text = re.sub(r'\s+', ' ', text).strip()
                        return text

                    df1_small['token_set'] = df1_small[match_col1].apply(normalize).str.split().apply(set)
                    df2_small['norm_match'] = df2_small[match_col2].apply(normalize)

                    tmp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
                    tmp_path = tmp_file.name
                    tmp_file.close()
                    pd.DataFrame(columns=include_cols1 + include_cols2).to_excel(tmp_path, index=False)

                    total_rows = len(df2_small)
                    progress_bar = st.progress(0)
                    status_text = st.empty()
                    start_time = time.time()

                    batch_size = 200
                    buffer_rows = []

                    for idx, row in df2_small.iterrows():
                        norm_val = row['norm_match']
                        if not norm_val:
                            continue
                        row_tokens = set(norm_val.split())
                        mask = df1_small['token_set'].apply(lambda x: row_tokens.issubset(x))
                        matched_rows = df1_small.loc[mask, include_cols1].copy()
                        if not matched_rows.empty:
                            for col in include_cols2:
                                matched_rows[col] = row[col]
                            buffer_rows.append(matched_rows)
                        if len(buffer_rows) >= batch_size:
                            batch_df = pd.concat(buffer_rows, ignore_index=True)
                            with pd.ExcelWriter(tmp_path, engine='openpyxl', mode='a',
                                                if_sheet_exists='overlay') as writer:
                                startrow = writer.sheets['Sheet1'].max_row
                                batch_df.to_excel(writer, index=False, header=False, startrow=startrow)
                            buffer_rows = []
                        progress_bar.progress((idx + 1) / total_rows)
                        status_text.text(f"Processing row {idx + 1}/{total_rows} ({(idx + 1) / total_rows * 100:.1f}%)")

                    if buffer_rows:
                        batch_df = pd.concat(buffer_rows, ignore_index=True)
                        with pd.ExcelWriter(tmp_path, engine='openpyxl', mode='a',
                                            if_sheet_exists='overlay') as writer:
                            startrow = writer.sheets['Sheet1'].max_row
                            batch_df.to_excel(writer, index=False, header=False, startrow=startrow)

                    end_time = time.time()
                    st.success(f"✅ Matching complete in {end_time - start_time:.2f} seconds")

                    preview_df = pd.read_excel(tmp_path, nrows=100)
                    st.subheader("Preview of Matched Results (first 100 rows)")
                    st.dataframe(preview_df)

                    with open(tmp_path, "rb") as f:
                        st.download_button("💾 Download Full Matched Results", data=f,
                                           file_name="matched_results.xlsx")

                    os.remove(tmp_path)

                except Exception as e:
                    st.error(f"❌ Error during matching: {e}")

    else:
        st.warning("⚠️ Please select at least 2 files for matching.")

# -----------------------------
# PART 2 - SEARCH & FILTER
# -----------------------------
st.header("🔹 Part 2: Search & Filter Data (Column-wise AND/OR + Global Keywords)")

uploaded_filter_file = st.file_uploader(
    "Upload an Excel file for filtering", type="xlsx", key="filter_file"
)

def filter_dataframe_columnwise_partial(df, column_keywords, logic="AND"):
    masks = []
    for col, keywords in column_keywords.items():
        col_series = df[col].astype(str).str.lower()
        keyword_masks = [col_series.str.contains(k.lower().strip(), regex=False, na=False) for k in keywords]
        if keyword_masks:
            masks.append(pd.concat(keyword_masks, axis=1).any(axis=1))
        else:
            masks.append(pd.Series([True]*len(df), index=df.index))
    if logic.upper() == "AND":
        final_mask = pd.concat(masks, axis=1).all(axis=1)
    else:
        final_mask = pd.concat(masks, axis=1).any(axis=1)
    return final_mask

if uploaded_filter_file:
    df_filter = pd.read_excel(uploaded_filter_file)
    st.success(f"✅ File {uploaded_filter_file.name} uploaded with {len(df_filter)} rows.")

    search_all = st.checkbox("🔎 Search across all columns (ignore column selection)")

    column_keywords = {}
    col_logic = "AND"
    keywords_input = ""
    global_logic = "OR"

    if not search_all:
        filter_cols = st.multiselect("Select columns to apply filters on", df_filter.columns.tolist())
        for col in filter_cols:
            keywords = st.text_input(f"Keywords for '{col}' (comma-separated)")
            if keywords:
                column_keywords[col] = [k.strip() for k in keywords.split(",") if k.strip()]
        col_logic_radio = st.radio(
            "Select cross-column logic for multiple columns",
            options=["AND (match all columns)", "OR (match any column)"],
            index=0
        )
        col_logic = "AND" if col_logic_radio.startswith("AND") else "OR"

    st.subheader("Global Search Options")
    keywords_input = st.text_input("Enter keywords to search across all columns (comma-separated)")
    global_logic_radio = st.radio(
        "Global keyword logic",
        options=["AND (match all keywords)", "OR (match any keyword)"],
        index=1
    )
    global_logic = "AND" if global_logic_radio.startswith("AND") else "OR"

    max_preview = st.number_input("Preview rows (max)", min_value=10, max_value=1000, value=200)

    if st.button("🔍 Apply Filter"):
        df_result = df_filter.copy()
        final_mask = pd.Series([True]*len(df_result), index=df_result.index)

        # Column-wise filtering
        if not search_all and column_keywords:
            col_mask = filter_dataframe_columnwise_partial(df_result, column_keywords, col_logic)
            final_mask &= col_mask

        # Global keywords filtering with AND/OR logic
        if keywords_input.strip():
            keywords = [k.lower().strip() for k in keywords_input.split(",") if k.strip()]
            masks = []
            for k in keywords:
                mask = df_result.astype(str).apply(lambda row: row.str.lower().str.contains(re.escape(k), na=False).any(), axis=1)
                masks.append(mask)
            if global_logic.upper() == "AND":
                global_mask = pd.concat(masks, axis=1).all(axis=1)
            else:
                global_mask = pd.concat(masks, axis=1).any(axis=1)
            final_mask &= global_mask

        df_result = df_result[final_mask]

        if df_result.empty:
            st.error("❌ No rows matched your filters.")
        else:
            st.success(f"✅ Found {len(df_result)} matching rows.")
            st.dataframe(df_result.head(max_preview))

            csv = df_result.to_csv(index=False).encode("utf-8")
            st.download_button("💾 Download Filtered Results (CSV)", data=csv,
                               file_name="filtered_results.csv")

            tmp_xlsx = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
            df_result.to_excel(tmp_xlsx.name, index=False)
            with open(tmp_xlsx.name, "rb") as f:
                st.download_button("💾 Download Filtered Results (XLSX)", data=f,
                                   file_name="filtered_results.xlsx")
            os.remove(tmp_xlsx.name)
# -----------------------------
# PART 3 - SEMANTIC MATCHING WITH ACTIVITIES
# -----------------------------
st.header("🔹 Part 3: Semantic Matching - Activities (Sentence Transformer)")

uploaded_part3_dc = st.file_uploader("Upload DC Log Excel file for Part 3", type="xlsx", key="part3_dc")
uploaded_part3_details = st.file_uploader("Upload Details Excel file for Part 3", type="xlsx", key="part3_details")

if uploaded_part3_dc and uploaded_part3_details:
    st.subheader("Select columns for Part 3")
    part3_dc_cols = pd.read_excel(uploaded_part3_dc, nrows=0).columns.tolist()
    part3_details_cols = pd.read_excel(uploaded_part3_details, nrows=0).columns.tolist()

    dc_doc_col = st.selectbox("Document Title column from DC Log", part3_dc_cols, index=1)  # default to "Title / Description"
    dc_number_col = st.selectbox("Document Number column from DC Log", part3_dc_cols, index=0)
    dc_status_col = st.selectbox("Status column from DC Log", part3_dc_cols, index=5)
    dc_function_col = st.selectbox("Function column from DC Log", part3_dc_cols, index=2)
    dc_discipline_col = st.selectbox("Discipline column from DC Log", part3_dc_cols, index=3)

    details_mh_col = st.selectbox("Start Structure column from Details", part3_details_cols, index=0)

    if st.button("Start Part 3 Matching"):
        st.info("⏳ Loading model and performing semantic matching... This may take time for large files.")
        try:
            import numpy as np
            from sentence_transformers import SentenceTransformer, util
            import time
            import tempfile

            # Load data
            dc_df = pd.read_excel(uploaded_part3_dc)
            details_df = pd.read_excel(uploaded_part3_details)

            # Load model
            model = SentenceTransformer("all-MiniLM-L6-v2")

            # Encode document titles
            doc_titles = dc_df[dc_doc_col].astype(str).tolist()
            doc_embeddings = model.encode(doc_titles, convert_to_tensor=True, show_progress_bar=True)

            # Define activities + synonyms
            activities = {
                "Excavation WIR": ["excavation", "excavating", "trenching", "digging"],
                "FDT WIR": ["fdt", "f.d.t", "field density test", "compaction test"],
                "Base WIR": ["blinding", "base concrete", "foundation base"],
                "Rings and Cover WIR": ["rings", "cover slab", "precast ring", "manhole cover"],
                "Cover WIR": ["cover", "slab cover"],
                "Backfilling WIR": ["backfilling", "backfill", "refilling"]
            }

            # Precompute embeddings
            activity_embeddings = {act: model.encode(keys, convert_to_tensor=True) for act, keys in activities.items()}

            results = []
            threshold = 0.70
            total_mh = len(details_df)
            progress_bar = st.progress(0)
            status_text = st.empty()
            start_time = time.time()

            for idx, mh_row in details_df.iterrows():
                mh = str(mh_row[details_mh_col])
                for act, act_emb in activity_embeddings.items():
                    sim = util.cos_sim(act_emb, doc_embeddings).cpu().numpy()
                    max_sim = np.max(sim, axis=0)
                    idxs = np.where(max_sim >= threshold)[0]
                    if len(idxs) > 0:
                        docs = dc_df.iloc[idxs]
                        results.append({
                            "Start Structure": mh,
                            "Activity": act,
                            "DocumentNo.": " / ".join(docs[dc_number_col].astype(str).tolist()),
                            "Title / Description": " / ".join(docs[dc_doc_col].astype(str).tolist()),
                            "Status": " / ".join(docs[dc_status_col].astype(str).tolist()),
                            "Function": " / ".join(docs[dc_function_col].astype(str).tolist()),
                            "Discipline": " / ".join(docs[dc_discipline_col].astype(str).tolist())
                        })
                    else:
                        results.append({
                            "Start Structure": mh,
                            "Activity": act,
                            "DocumentNo.": "Not Found",
                            "Title / Description": "Not Found",
                            "Status": "Not Found",
                            "Function": "Not Found",
                            "Discipline": "Not Found"
                        })
                progress_bar.progress((idx+1)/total_mh)
                status_text.text(f"Processing Start Structure {idx+1}/{total_mh} ({(idx+1)/total_mh*100:.1f}%)")

            final_df = pd.DataFrame(results)
            end_time = time.time()
            st.success(f"✅ Part 3 Matching completed in {end_time-start_time:.2f} seconds")

            st.subheader("Preview first 100 rows")
            st.dataframe(final_df.head(100))

            tmp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
            final_df.to_excel(tmp_file.name, index=False)
            with open(tmp_file.name, "rb") as f:
                st.download_button("💾 Download Part 3 Results (XLSX)", data=f,
                                   file_name="activities_matched_merged.xlsx")
            os.remove(tmp_file.name)

        except Exception as e:
            st.error(f"❌ Error in Part 3: {e}")
